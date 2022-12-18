import openpyxl

filename = "Results.xlsx"
wb = openpyxl.load_workbook(filename)
ws = wb.worksheets[0]
index = 1

def write_single_chart_dict(title: str, labels: list, keys: list, values: dict):
    global index

    ws.cell(index, 1, title)

    for compt in range(len(labels)):
        pos = compt + 2
        ws.cell(index, pos, labels[compt])
        ws.cell(index + 1, pos, values.setdefault(keys[compt], 0))

    index += 2

# Wished structure
# TITLE   LABEL 1 LABEL 2 LABEL 3
# SERIE 1 DATA 11 DATA 12 DATA 13
# SERIE 2 DATA 21 DATA 22 DATA 23
# SERIE 3 DATA 31 DATA 32 DATA 33
def write_chart_dict(title: str, series_labels: list, labels: list, series: list, keys: list, values: dict):
    global index

    ws.cell(index, 1, title)

    for compt in range(len(labels)):
        pos = compt + 2
        ws.cell(index, pos, labels[compt])

    index += 1

    for compt in range(len(series)):
        ws.cell(index, 1, series_labels[compt])

        for compt2 in range(len(labels)):
            pos = compt2 + 2
            ws.cell(index, pos, values.setdefault(series[compt], {}).setdefault(keys[compt2], 0))

        index += 1

# Save the file
def save_file():
    wb.save(filename)

def close():
    wb.close()